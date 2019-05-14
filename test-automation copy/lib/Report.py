import os, re, sys
import traceback
import subprocess
import datetime
import time
from time import strftime
import xmlrpclib
import fnmatch, csv
from Trace import *
import HTML
import csv
from openpyxl import *
from openpyxl.styles import *
from openpyxl.chart import (AreaChart,LineChart,Reference,Series,)
from email import Header

class HtmlReport:
    def __init__(self,htmlFile,logDir = None):
        if logDir == None:
            LogDir = os.path.join(os.path.abspath(os.path.join(os.path.dirname(os.path.abspath("Report.py")), os.pardir)), "logs")
            if  not os.path.exists(LogDir):
                os.mkdir(LogDir)
        else:
            LogDir = logDir
            if  not os.path.exists(LogDir):
                os.mkdir(LogDir)
        self.htmlFile = os.path.join(LogDir,htmlFile+"_"+strftime("%Y-%m-%d_%H-%M-%S")+".html")
        self.htmlFileHandler = None
        self.htmlTitle = ""
        self.htmlTable = None
        self.htmlLink = ""
        self.htmlHeader = ""
        self.html = ""
        self.trace = Trace("Report.py")
    
    def openHtml(self):
        try:
            if self.htmlFileHandler == None:
                self.htmlFileHandler = open(self.htmlFile,"w")
            else:
                raise Exception("The file:%s is already opened."%self.htmlFile)
        except:
            self.trace.Error("ExceptionType:"+str(sys.exc_info()[0]))
            self.trace.Error("ExceptionDetails:"+str(sys.exc_info()[1]))
            self.trace.Error("TraceBack:"+str(traceback.extract_tb(sys.exc_info()[2])))
            raise
    
    def createHtmlTableHeader(self,header):
        try:
            if self.htmlTable == None:
                self.htmlTable = HTML.Table(header_row=header)
        except:
            self.trace.Error("ExceptionType:"+str(sys.exc_info()[0]))
            self.trace.Error("ExceptionDetails:"+str(sys.exc_info()[1]))
            self.trace.Error("TraceBack:"+str(traceback.extract_tb(sys.exc_info()[2])))
            raise
    
    def addRowToHtml(self,row,color='Lime'):
        try:
            newRow = HTML.TableRow(row, bgcolor=color)
            self.htmlTable.rows.append(newRow)
        except:
            self.trace.Error("ExceptionType:"+str(sys.exc_info()[0]))
            self.trace.Error("ExceptionDetails:"+str(sys.exc_info()[1]))
            self.trace.Error("TraceBack:"+str(traceback.extract_tb(sys.exc_info()[2])))
            raise
    
    def HtmlHeader(self,text,value):
        try:
            self.htmlHeader = self.htmlHeader + str(HTML.Header(text, value)) + '<p>\n'
        except:
            self.trace.Error("ExceptionType:"+str(sys.exc_info()[0]))
            self.trace.Error("ExceptionDetails:"+str(sys.exc_info()[1]))
            self.trace.Error("TraceBack:"+str(traceback.extract_tb(sys.exc_info()[2])))
            raise
        
    def HtmlLink(self,text,link):
        try:
            self.htmlLink = self.htmlLink + str(HTML.link(text, link)) + '<p>\n'
        except:
            self.trace.Error("ExceptionType:"+str(sys.exc_info()[0]))
            self.trace.Error("ExceptionDetails:"+str(sys.exc_info()[1]))
            self.trace.Error("TraceBack:"+str(traceback.extract_tb(sys.exc_info()[2])))
            raise
    
    def HtmlTitle(self,text):
        try:
            self.htmlTitle = self.htmlTitle + str(HTML.Title(text)) + '<p>\n'
        except:
            self.trace.Error("ExceptionType:"+str(sys.exc_info()[0]))
            self.trace.Error("ExceptionDetails:"+str(sys.exc_info()[1]))
            self.trace.Error("TraceBack:"+str(traceback.extract_tb(sys.exc_info()[2])))
            raise
        
    def getTheHtmlCode(self):
        try:
            if self.htmlTitle != "":
                self.html = self.html + self.htmlTitle
            if self.htmlHeader != "":
                self.html = self.html + self.htmlHeader
            if str(self.htmlTable) != "":
                self.html = self.html + str(self.htmlTable)
            if self.htmlLink != "":
                self.html = self.html + self.htmlLink
            return str(self.html)
        except:
            self.trace.Error("ExceptionType:"+str(sys.exc_info()[0]))
            self.trace.Error("ExceptionDetails:"+str(sys.exc_info()[1]))
            self.trace.Error("TraceBack:"+str(traceback.extract_tb(sys.exc_info()[2])))
            raise
    
    def writeToHtml(self):
        try:
            htmlCode = self.getTheHtmlCode()
            if htmlCode != "":
                self.htmlFileHandler.write(htmlCode+'<p>\n')
#             else:
#                 raise Exception("There is not Html String to write.The Html String is:%s"%htmlCode)
        except:
            self.trace.Error("ExceptionType:"+str(sys.exc_info()[0]))
            self.trace.Error("ExceptionDetails:"+str(sys.exc_info()[1]))
            self.trace.Error("TraceBack:"+str(traceback.extract_tb(sys.exc_info()[2])))
            raise
        
    def getHhtmlFilePath(self):
        return self.htmlFile 
                    
    def __del__(self):
        try:
            #self.writeToHtml()
            print "deleting %s file"%self.htmlFile
            self.htmlFileHandler.close()
        except:
            self.trace.Error("ExceptionType:"+str(sys.exc_info()[0]))
            self.trace.Error("ExceptionDetails:"+str(sys.exc_info()[1]))
            self.trace.Error("TraceBack:"+str(traceback.extract_tb(sys.exc_info()[2])))
            raise
        
class CsvReport:
    def __init__(self,csvFile,logDir = None):
        if logDir == None:
            LogDir = os.path.join(os.path.abspath(os.path.join(os.path.dirname(os.path.abspath("Report.py")), os.pardir)), "logs")
            if  not os.path.exists(LogDir):
                os.mkdir(LogDir)
        else:
            LogDir = logDir
            if  not os.path.exists(LogDir):
                os.mkdir(LogDir)
        self.csvFile = os.path.join(LogDir,csvFile+"_"+strftime("%Y-%m-%d_%H-%M-%S")+".csv")
        self.csv = None
        self.csvFileHandler = None
        self.trace = Trace("Report.py")
            
    def openCsv(self):
        try:
            if self.csvFileHandler == None:
                self.csvFileHandler = open(self.csvFile,"wb")
            else:
                raise Exception("The file:%s is already opened."%self.csvFile)
        except:
            self.trace.Error("ExceptionType:"+str(sys.exc_info()[0]))
            self.trace.Error("ExceptionDetails:"+str(sys.exc_info()[1]))
            self.trace.Error("TraceBack:"+str(traceback.extract_tb(sys.exc_info()[2])))
            raise
         
    def addRowToCsv(self,row):
        try:
            if self.csv == None:
                self.csv = csv.writer(self.csvFileHandler)
                self.csv.writerow(row)
            else:
                self.csv.writerow(row)
        except:
            self.trace.Error("ExceptionType:"+str(sys.exc_info()[0]))
            self.trace.Error("ExceptionDetails:"+str(sys.exc_info()[1]))
            self.trace.Error("TraceBack:"+str(traceback.extract_tb(sys.exc_info()[2])))
            raise
    
    def getCsvFilePath(self):
        return self.csvFile 
      
    def __del__(self):
        try:
            print "deleting %s file"%self.csvFile
            self.csvFileHandler.close()
        except:
            self.trace.Error("ExceptionType:"+str(sys.exc_info()[0]))
            self.trace.Error("ExceptionDetails:"+str(sys.exc_info()[1]))
            self.trace.Error("TraceBack:"+str(traceback.extract_tb(sys.exc_info()[2])))
            raise
        
class XlsReport:
    def __init__(self,xlsFile,logDir = None):
        if logDir == None:
            LogDir = os.path.join(os.path.abspath(os.path.join(os.path.dirname(os.path.abspath("Report.py")), os.pardir)), "logs")
            if  not os.path.exists(LogDir):
                os.mkdir(LogDir)
        else:
            LogDir = logDir
            if  not os.path.exists(LogDir):
                os.mkdir(LogDir)
        self.xlsFile = os.path.join(LogDir,xlsFile+"_"+strftime("%Y-%m-%d_%H-%M-%S")+".xlsx")
        self.wrkBook = Workbook()
        self.wrkBook.save(self.xlsFile)
        #wrkBook.close()
        self.wrkSheetCurIndex=1
        self.curRowIndex={}
        self.header = {}
        self.trace = Trace("Report.py")
            
    def openWrkBook(self):        
        if not os.path.exists(self.xlsFile):
            raise Exception("File not found: %s"%self.xlsFile)
        wrkBook = load_workbook(self.xlsFile)
        if wrkBook is None:
            raise Exception("Cannot open Excel file : %s"%xlsFile)
        return wrkBook
    
    def getColumnAlphabetFromIndex(self,index):
        return str(unichr(65+int(index)))
        
    def getColumnAlphabetFromName(self,wrkSheetName , colName):
        return self.getColumnAlphabetFromIndex(self.header[wrkSheetName].index(colName))
        
    def createWrkSheet(self , wrkSheetName , Header,index=None):
        if wrkSheetName is None:
            self.trace.Error("Argument sheet name is not provided")
            raise Exception("Sheet name not provided")
        try:
            wrkBook = self.openWrkBook()
            try:
                wrkSheet = wrkBook.get_sheet_by_name(wrkSheetName)
            except KeyError:
                wrkSheet = self.addWrkSheet(wrkBook , wrkSheetName)
            count=0
            for each in Header:
                loc = self.getColumnAlphabetFromIndex(count)+"1"
                wrkSheet[loc] = each
                count+=1
            wrkBook.save(self.xlsFile)
            self.wrkSheetCurIndex += 1
            self.curRowIndex[wrkSheetName]=2
            self.header[wrkSheetName]=Header
            #wrkBook.close()
        except:
            self.trace.Error("ExceptionType:"+str(sys.exc_info()[0]))
            self.trace.Error("ExceptionDetails:"+str(sys.exc_info()[1]))
            self.trace.Error("TraceBack:"+str(traceback.extract_tb(sys.exc_info()[2])))
            raise
        
        
    def writeLineToWrkSheet(self , wrkSheetName , row):
        if wrkSheetName is None:
            self.trace.Error("Argument sheet name is not provided")
            raise Exception("Sheet name not provided")
        try:
            wrkBook = self.openWrkBook()
            try:
                wrkSheet = wrkBook.get_sheet_by_name(wrkSheetName)
            except KeyError:
                wrkSheet = self.addWrkSheet(wrkBook , wrkSheetName)
            count=0
            for each in row:
                loc = self.getColumnAlphabetFromName(each)+str(self.curRowIndex)
                wrkSheet[loc] = each
                count+=1
                self.curRowIndex[wrkSheetName] += 1
            wrkBook.save(self.xlsFile)
            #wrkBook.close()
        except:
            self.trace.Error("ExceptionType:"+str(sys.exc_info()[0]))
            self.trace.Error("ExceptionDetails:"+str(sys.exc_info()[1]))
            self.trace.Error("TraceBack:"+str(traceback.extract_tb(sys.exc_info()[2])))
            raise
        
    def writeLineToWrkSheetWithOptions(self , wrkSheetName , row=None):
        colorDict = { "RED":colors.RED , "BLUE":colors.BLUE , "GREEN":colors.GREEN , "YELLOW":colors.YELLOW }
        if wrkSheetName is None:
            self.trace.Error("Argument sheet name is not provided")
            raise Exception("Sheet name not provided")
        try:
            wrkBook = self.openWrkBook()
            try:
                wrkSheet = wrkBook.get_sheet_by_name(wrkSheetName)
            except KeyError:
                wrkSheet = self.addWrkSheet(wrkBook , wrkSheetName)
            count=0
            for value in row:
                loc = self.getColumnAlphabetFromIndex(count)+str(self.curRowIndex[wrkSheetName])
                if isinstance(value, list) or isinstance(value, tuple):
                    wrkSheet[loc] = value[0]
                    if value[1] is not None:
                        ft = Font(color=colorDict[value[1]])
                        wrkSheet[loc].font = ft
                else:
                    wrkSheet[loc]=value
                count+=1
            self.curRowIndex[wrkSheetName] += 1
            wrkBook.save(self.xlsFile)
            #wrkBook.close()
        except:
            self.trace.Error("ExceptionType:"+str(sys.exc_info()[0]))
            self.trace.Error("ExceptionDetails:"+str(sys.exc_info()[1]))
            self.trace.Error("TraceBack:"+str(traceback.extract_tb(sys.exc_info()[2])))
            raise
        
    def writeCellToWrkSheet(self , wrkBook , wrkSheetName , loc , data , cellProperties=None):
        colorDict = { "RED":colors.RED , "BLUE":colors.BLUE , "GREEN":colors.GREEN , "YELLOW":colors.YELLOW }
        if wrkSheetName is None:
            self.trace.Error("Argument sheet name is not provided")
            raise Exception("Sheet name not provided")
        try:
            try:
                wrkSheet = wrkBook.get_sheet_by_name(wrkSheetName)
            except KeyError:
                wrkSheet = self.addWrkSheet(wrkBook , wrkSheetName)
            if cellProperties is not None:
                if cellProperties.has_key("color"):
                    ft = Font(color=colorDict[cellProperties['color']])
                    wrkSheet[loc].font = ft
            wrkSheet[loc] = data
            #wrkBook.save(self.xlsFile)
            #wrkBook.close()
        except:
            self.trace.Error("ExceptionType:"+str(sys.exc_info()[0]))
            self.trace.Error("ExceptionDetails:"+str(sys.exc_info()[1]))
            self.trace.Error("TraceBack:"+str(traceback.extract_tb(sys.exc_info()[2])))
            raise
         
    def getRowCursorFromSheet(self , wrkSheet):
        return self.curRowIndex[wrkSheet]
         
    def addWrkSheet(self , wrkBook , sheetName=None):
        if sheetName is None:
            self.trace.Error("Argument sheet name is not provided")
            raise Exception("Sheet name not provided")
        try:
            sheets = wrkBook.get_sheet_names()
            if 'Sheet' in sheets and len(sheets) == 1:
                wrkSheet = wrkBook.active
                wrkSheet.title=sheetName
            else:
                wrkSheet = wrkBook.create_sheet(title=sheetName)  
            return wrkSheet
        except:
            self.trace.Error("ExceptionType:"+str(sys.exc_info()[0]))
            self.trace.Error("ExceptionDetails:"+str(sys.exc_info()[1]))
            self.trace.Error("TraceBack:"+str(traceback.extract_tb(sys.exc_info()[2]))) 
            raise Exception("Error occured in adding work sheet")
        
                
    def createLineChart(self ,chartSheet , referenceSheet , dataSheets  , ChartDetails=None):
        '''
        
        '''
        try:
            wrkBook = self.openWrkBook()
            try:
                wrkSheet = wrkBook.get_sheet_by_name(referenceSheet)
            except KeyError:
                print 'Reference Sheet does not exists'
                raise
            chart = LineChart()
            chart.title = ChartDetails['title']
            chart.style = ChartDetails['style']
            chart.x_axis.title = ChartDetails['x-axis-title']
            chart.y_axis.title = ChartDetails['y-axis-title']
            maxRow = ChartDetails['maxRow']
            maxColumn = ChartDetails['maxColumn']
            cats = Reference(wrkSheet, min_col=1, min_row=1, max_row=maxRow)
            chart.set_categories(cats)
            for sheet in dataSheets:
                try:
                    wrkSheet = wrkBook.get_sheet_by_name(sheet)
                except KeyError:
                    print 'Data Sheet [%s] does not exists'%sheet
                    raise
                data = Reference(wrkSheet, min_col=2, min_row=1, max_col=maxColumn, max_row=maxRow)
                chart.add_data(data, titles_from_data=True)
            #self.createWrkSheet("Chart", [""])
            wrkBook = self.openWrkBook()
            #print wrkBook.get_sheet_names()
            chartSheet = wrkBook.get_sheet_by_name(chartSheet)
            chartLoc = 'K10'
            chartSheet.add_chart(chart, chartLoc)
            wrkBook.save(self.xlsFile)  
        except Exception as e:
            traceback.print_exc(e)
            self.trace.Error("ExceptionType:"+str(sys.exc_info()[0]))
            self.trace.Error("ExceptionDetails:"+str(sys.exc_info()[1]))
            self.trace.Error("TraceBack:"+str(traceback.extract_tb(sys.exc_info()[2]))) 
            raise Exception("Error occured in adding work sheet")

    def createGraphFromSingleColumns(self ,ColumnAlphabet):
        #wkBook = load_workbook("C:\\testxl\\WeMoStatLive_2015-11-23_23-33-08.xlsx")        
        #self.xlsFile = "C:\\Users\\imran.ali\\workspace\\testAutomation_sanity\\logs\\WeMoStatLive_2015-11-27_12-35-18.xlsx"
        wkBook = self.openWrkBook()
        datasheets=[]
        for sheet in wkBook.get_sheet_names():
            if 'Chart' not in sheet:
                datasheets.append(sheet)
        sheetCount=0
        dataDict = {}
        highestRow=0
        highestCol=0
        for sheet in datasheets:
            wkSheet = wkBook.get_sheet_by_name(sheet)
            data_range = ColumnAlphabet + "2:" + ColumnAlphabet + str(wkSheet.max_row)
            dataList = [] 
            for row in wkSheet.iter_rows(data_range):
                for cell in row:
                    dataList.append(cell.value)
            dataDict[sheet]=dataList
            if wkSheet.max_row > highestRow:
                highestRow=wkSheet.max_row
            if wkSheet.max_column > highestCol:
                highestCol=wkSheet.max_column
        columnName = wkSheet["%s1"%ColumnAlphabet].value
        chartSheetName = columnName + "Chart"
        self.createWrkSheet(chartSheetName, ["Iteration"]+datasheets)
        wkBook = self.openWrkBook()
        for i in range(2,highestRow+1):
            loc = "%s%s"%(self.getColumnAlphabetFromIndex(0),i) 
            self.writeCellToWrkSheet(wkBook,chartSheetName, loc, i-1)
        sheetCount=1
        wkBook.save(self.xlsFile)
        for sheet in datasheets:
            count=2
            for elem in dataDict[sheet]:
                loc = "%s%s"%(self.getColumnAlphabetFromIndex(sheetCount),count) 
                self.writeCellToWrkSheet(wkBook,chartSheetName, loc, elem)
                count+=1
            #self.writeLineToWrkSheetWithOptions(columnName, [data]+dataDict[data])
            sheetCount+=1
        wkBook.save(self.xlsFile)
        chartDetails={}
        chartDetails['title']=chartSheetName
        chartDetails['style']=13
        chartDetails['x-axis-title']="Iteration"
        chartDetails['y-axis-title']=columnName
        chartDetails['maxRow']=highestRow
        chartDetails['maxColumn']=len(datasheets)+1
        self.createLineChart( chartSheetName,chartSheetName, [chartSheetName],chartDetails)
        
    def copyAndCreateGraph(self , fileName , ColumnAlphabet):
        saveFile = self.xlsFile
        wkBook = load_workbook(self.xlsFile)
        LogDir = os.path.join(os.path.abspath(os.path.join(os.path.dirname(os.path.abspath("Report.py")), os.pardir)), "logs")
        fileName = os.path.join(LogDir,fileName)
        wkBook.save(fileName)
        self.xlsFile = fileName
        wkBook = self.openWrkBook()
        datasheets=[]
        for sheet in wkBook.get_sheet_names():
            if 'Chart' not in sheet:
                datasheets.append(sheet)
        sheetCount=0
        dataDict = {}
        highestRow=0
        highestCol=0
        for sheet in datasheets:
            wkSheet = wkBook.get_sheet_by_name(sheet)
            data_range = ColumnAlphabet + "2:" + ColumnAlphabet + str(wkSheet.max_row)
            dataList = [] 
            for row in wkSheet.iter_rows(data_range):
                for cell in row:
                    dataList.append(cell.value)
            dataDict[sheet]=dataList
            if wkSheet.max_row > highestRow:
                highestRow=wkSheet.max_row
            if wkSheet.max_column > highestCol:
                highestCol=wkSheet.max_column
        columnName = wkSheet["%s1"%ColumnAlphabet].value
        chartSheetName = columnName + "Chart"
        self.createWrkSheet(chartSheetName, ["Iteration"]+datasheets)
        wkBook = self.openWrkBook()
        for i in range(2,highestRow+1):
            loc = "%s%s"%(self.getColumnAlphabetFromIndex(0),i) 
            self.writeCellToWrkSheet(wkBook,chartSheetName, loc, i-1)
        sheetCount=1
        wkBook.save(self.xlsFile)
        for sheet in datasheets:
            count=2
            for elem in dataDict[sheet]:
                loc = "%s%s"%(self.getColumnAlphabetFromIndex(sheetCount),count) 
                self.writeCellToWrkSheet(wkBook,chartSheetName, loc, elem)
                count+=1
            #self.writeLineToWrkSheetWithOptions(columnName, [data]+dataDict[data])
            sheetCount+=1
        wkBook.save(self.xlsFile)
        chartDetails={}
        chartDetails['title']=chartSheetName
        chartDetails['style']=20
        chartDetails['x-axis-title']="Iteration"
        chartDetails['y-axis-title']=columnName
        chartDetails['maxRow']=highestRow
        chartDetails['maxColumn']=len(datasheets)+1
        self.createLineChart( chartSheetName,chartSheetName, [chartSheetName],chartDetails)
        self.xlsFile=saveFile
        
    def createGraphFromData(self):
        #wkBook = load_workbook("C:\\testxl\\WeMoStatLive_2015-11-23_23-33-08.xlsx")
        #self.xlsFile = "C:\\Users\\imran.ali\\workspace\\testAutomation_sanity\\logs\\WeMoStatLive_2015-11-24_19-28-10.xlsx"
        wkBook = self.openWrkBook()
        datasheets = wkBook.get_sheet_names()
        sheetCount=0
        dataDict = {}
        for sheet in datasheets:
            wkSheet = wkBook.get_sheet_by_name(sheet)
            for row in wkSheet.iter_rows("F2:F%s"%wkSheet.max_row):
                for cell in row:
                    print cell.value
            maxRow = wkSheet.max_row+1
            maxCol = 8
            Count=1
            for i in range(2,maxRow):
                print Count
                iterationCount = int(wkSheet.cell(row=i , column=1).value)
                print iterationCount
                if sheetCount == 0:
                    if iterationCount != Count:
                        val = [-1]
                        dataDict[Count]=val
                        val = [wkSheet.cell(row=i , column=6).value]
                        dataDict[iterationCount]=val
                    else:
                        val = [wkSheet.cell(row=i , column=6).value]
                        dataDict[iterationCount]=val
                else:
                    #print dataList[0]
                    listVal = dataDict[iterationCount]
                    listVal.append( wkSheet.cell(row=i , column=6).value)
                    dataDict[iterationCount]=listVal
#                 if sheetCount == 0:
#                     dataList.append(val)
#                 else:
#                     dataList[iterationCount]=val
                Count += 1
                print dataDict
            sheetCount += 1
        self.xlsFile="C:\\testxl\\WeMoStatLive_2015-11-23_23-33-08.xlsx"
        self.createWrkSheet("MemFree", ["iteration"]+datasheets)
        for data in dataDict:
            self.writeLineToWrkSheetWithOptions("MemFree", [data]+dataDict[data])
        chartDetails={}
        chartDetails['title']="MemFree"
        chartDetails['style']=20
        chartDetails['x-axis-title']="Iteration"
        chartDetails['y-axis-title']="MemFree"
        chartDetails['maxRow']=6
        chartDetails['maxColumn']=3
        r.createLineChart( "MemFree","MemFree", ["MemFree"],chartDetails)

class XLSReportParser():

    def __init__(self,xls_report_file):
        self.report = load_workbook(xls_report_file)
        self.sheetnames = self.report.get_sheet_names()

    def getColumnTuple(self,sheet,coloumnIndex):
        try:
            data = []
            sheetname = sheet
            xlssheet = self.report.get_sheet_by_name(sheetname)
            coloumn = xlssheet.columns[coloumnIndex]
            for cell in coloumn:
                data.append(cell.value)
            return data
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            raise Exception(str(exc_obj.message), fname, exc_tb.tb_lineno,locals())

    def getRowTuple(self,sheet,rowIndex):
        try:
            data = []
            sheetname = sheet
            xlssheet = self.report.get_sheet_by_name(sheetname)
            row = xlssheet.rows[rowIndex]
            for cell in row:
                data.append(cell.value)
            return data
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            raise Exception(str(exc_obj.message), fname, exc_tb.tb_lineno,locals())


if __name__ == "__main__":
    #createGraphFromData()
    r = XlsReport("testFile")
    r.createGraphFromSingleColumns("F")
    r.copyAndCreateGraph("C:\\Users\\imran.ali\\workspace\\testAutomation_sanity\\logs\\AnotherFile.xlsx" , "G")
    r.copyAndCreateGraph("C:\\Users\\imran.ali\\workspace\\testAutomation_sanity\\logs\\AnotherFile2.xlsx" , "H")
    #r.createGraphFromSingleColumns("E")
# #     for i in range(10):
#     r.createWrkSheet("Sheet1", ["Number","Batch1","Batch2","Batch3","Batch4"])
#     dataList = [[1, 60, 40,22,33],
#                     [2, 50, 65, 21, 77],
#                     [3, 20, 20 , 87 , 99],
#                     [4, 50, 40 , 32 , 65],
#                     [5, 65, 55 , 98 , 55],
#                     [6, 90, 20,66,44],
#                     ]
#     for data in dataList:
#             r.writeLineToWrkSheetWithOptions("Sheet1", data)
#     r.createWrkSheet("Sheet2", ["Number","Batch5","Batch6"])
#     dataList1 = [[1, 40, 30],
#                     [2, 40, 25],
#                     [3, 50, 30],
#                     [4, 30, 10],
#                     [5, 25, 5],
#                     [6, 50, 10],
#                     ]
#     for data in dataList1:
#             r.writeLineToWrkSheetWithOptions("Sheet2", data)
#     chartDetails={}
#     chartDetails['title']="Imran"
#     chartDetails['style']=20
#     chartDetails['x-axis-title']="Imran"
#     chartDetails['y-axis-title']="Ali"
#     chartDetails['maxRow']=6
#     chartDetails['maxColumn']=5
#     r.createLineChart( "Sheet1", ["Sheet1","Sheet2"],chartDetails)
#                 
#         
#     #r.writeLineToWrkSheetWithOptions("2", {"a":{'color':'RED'},"b":{'color':'RED'},"c":{'color':'RED'},"d":{'color':'RED'}})
#     
#     #r.writeCellToWrkSheet("test2", "C4", "data" , {'color':'RED','bold':True})